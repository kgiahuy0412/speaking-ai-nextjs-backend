import json
import sys
from collections import Counter
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

sys.stdout.reconfigure(encoding="utf-8")


INPUT = Path(r"C:\Users\DELL\Downloads\tong_hop_check_am_thanh_tieu_chuan.xlsx")


def clean(value):
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return value


workbook = load_workbook(INPUT, read_only=False, data_only=False)
result = {
    "file": str(INPUT),
    "size_bytes": INPUT.stat().st_size,
    "sheets": [],
}

for sheet in workbook.worksheets:
    nonempty_rows = []
    value_counts = Counter()
    for row in sheet.iter_rows():
        values = [clean(cell.value) for cell in row]
        if any(value not in (None, "") for value in values):
            while values and values[-1] in (None, ""):
                values.pop()
            nonempty_rows.append({"row": row[0].row, "values": values})
            for value in values:
                if isinstance(value, str) and value.strip():
                    value_counts[value.strip()] += 1

    result["sheets"].append(
        {
            "title": sheet.title,
            "state": sheet.sheet_state,
            "max_row": sheet.max_row,
            "max_column": sheet.max_column,
            "merged_ranges": [str(item) for item in sheet.merged_cells.ranges],
            "auto_filter": str(sheet.auto_filter.ref) if sheet.auto_filter.ref else None,
            "freeze_panes": str(sheet.freeze_panes) if sheet.freeze_panes else None,
            "nonempty_row_count": len(nonempty_rows),
            "rows": nonempty_rows,
            "common_text": value_counts.most_common(30),
        }
    )

print(json.dumps(result, ensure_ascii=False, indent=2))
